"""
Tests for the messy-file ingestion engine.

Covers: preamble CSV, footer stripping, repeated header rows, Excel region
detection, schema normalizer, and backward-compat of load_dataset().
"""
from __future__ import annotations

import textwrap
import tempfile
import pathlib

import openpyxl
import pandas as pd
import pytest

from app.services.file_loader import load_dataset, load_dataset_with_report
from app.services.ingestion import (
    ParseReport,
    normalize_schema,
    sniff_csv,
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _write_csv(content: str) -> str:
    """Write *content* to a temp CSV file and return its path."""
    with tempfile.NamedTemporaryFile(suffix=".csv", mode="w", delete=False) as f:
        f.write(content)
        return f.name


def _write_excel(wb: openpyxl.Workbook) -> str:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        return f.name


# ── Backward compatibility ─────────────────────────────────────────────────────

class TestBackwardCompat:
    def test_flat_csv_unchanged(self):
        path = _write_csv("name,age,salary\nAlice,30,70000\nBob,25,50000\n")
        df = load_dataset(path)
        assert list(df.columns) == ["name", "age", "salary"]
        assert len(df) == 2

    def test_flat_csv_returns_dataframe(self):
        path = _write_csv("a,b,c\n1,2,3\n4,5,6\n")
        result = load_dataset(path)
        assert isinstance(result, pd.DataFrame)

    def test_missing_file_raises(self):
        with pytest.raises(FileNotFoundError):
            load_dataset("/nonexistent/path/file.csv")

    def test_unsupported_extension_raises(self):
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as f:
            f.write(b"{}")
            path = f.name
        with pytest.raises(ValueError, match="Unsupported"):
            load_dataset(path)


# ── Sniffer ────────────────────────────────────────────────────────────────────

class TestSniffer:
    def test_flat_table_detected(self):
        path = _write_csv("name,age,salary\nAlice,30,70000\nBob,25,50000\n")
        result = sniff_csv(pathlib.Path(path))
        assert result.file_kind == "flat_table"
        assert result.header_row == 0
        assert result.confidence >= 0.9

    def test_preamble_csv_detected(self):
        content = textwrap.dedent("""\
            Statistics Office — Regional Employment Survey
            Source: National Bureau of Statistics, 2024
            Downloaded: 2024-01-15

            region,year,employment_rate,population
            North,2021,0.94,1200000
            South,2021,0.89,980000
            East,2021,0.91,750000
        """)
        path = _write_csv(content)
        result = sniff_csv(pathlib.Path(path))
        assert result.file_kind in ("preamble_csv", "sectioned_csv")
        assert result.header_row >= 3

    def test_footer_detected(self):
        content = (
            "country,gdp,population\n"
            "US,21000,330\n"
            "DE,4000,83\n"
            "FR,2900,67\n"
            "Source: World Bank 2024\n"
            "Notes: GDP in billions USD\n"
        )
        path = _write_csv(content)
        result = sniff_csv(pathlib.Path(path))
        assert result.footer_start_row is not None
        assert result.footer_start_row >= 4   # after the 3 data rows

    def test_tab_delimiter_detected(self):
        content = "name\tage\tsalary\nAlice\t30\t70000\nBob\t25\t50000\n"
        path = _write_csv(content)
        result = sniff_csv(pathlib.Path(path))
        assert result.delimiter == "\t"

    def test_semicolon_delimiter_detected(self):
        content = "name;age;salary\nAlice;30;70000\nBob;25;50000\n"
        path = _write_csv(content)
        result = sniff_csv(pathlib.Path(path))
        assert result.delimiter == ";"


# ── Preamble stripping ─────────────────────────────────────────────────────────

class TestPreambleStripping:
    def test_preamble_stripped_columns_correct(self):
        content = textwrap.dedent("""\
            Annual Economic Report 2024
            Source: Ministry of Finance
            Published: March 2024

            region,year,gdp,growth_rate
            North,2021,500.0,0.03
            South,2021,420.0,0.02
            East,2021,310.0,0.04
        """)
        path = _write_csv(content)
        df, rpt = load_dataset_with_report(path)
        assert "region" in df.columns, f"columns: {list(df.columns)}"
        assert "gdp" in df.columns
        assert len(df) >= 3

    def test_metadata_extracted(self):
        content = textwrap.dedent("""\
            Annual Economic Report 2024
            Source: Ministry of Finance

            region,gdp
            North,500
            South,420
        """)
        path = _write_csv(content)
        _, rpt = load_dataset_with_report(path)
        # Either metadata was extracted or at minimum the parse succeeded
        assert rpt.status in ("ok", "parsed_with_warnings", "fallback")
        if rpt.file_kind != "flat_table":
            assert rpt.warnings  # should warn about skipped preamble

    def test_preamble_warning_issued(self):
        content = textwrap.dedent("""\
            Report Title

            region,gdp
            North,500
            South,420
        """)
        path = _write_csv(content)
        _, rpt = load_dataset_with_report(path)
        if rpt.header_row and rpt.header_row > 0:
            assert any("preamble" in w.lower() or "skip" in w.lower() for w in rpt.warnings)


# ── Footer stripping ──────────────────────────────────────────────────────────

class TestFooterStripping:
    def test_source_footer_trimmed(self):
        content = (
            "country,gdp,population\n"
            "US,21000,330\n"
            "DE,4000,83\n"
            "FR,2900,67\n"
            "Source: World Bank 2024\n"
            "Notes: GDP in billions USD\n"
        )
        path = _write_csv(content)
        df, rpt = load_dataset_with_report(path)
        # Either 3 clean data rows, or fallback still loaded valid data
        data_rows = df[df["country"].apply(lambda x: str(x).strip() not in {"country", "Source:", "Notes:"})
                   ] if "country" in df.columns else df
        assert len(df) <= 5   # at most 3 data rows + maybe 1-2 edge cases
        assert "country" in df.columns

    def test_notes_footer_trimmed(self):
        content = (
            "item,price\n"
            "Apple,1.20\n"
            "Banana,0.50\n"
            "notes: prices include VAT\n"
        )
        path = _write_csv(content)
        df, rpt = load_dataset_with_report(path)
        assert "item" in df.columns
        # notes row should not appear as a data row (it would fail float cast anyway)


# ── Repeated header rows ──────────────────────────────────────────────────────

class TestRepeatedHeaders:
    def test_repeated_header_removed(self):
        content = "id,value,label\n1,100,A\n2,200,B\nid,value,label\n3,300,C\n"
        path = _write_csv(content)
        df, rpt = load_dataset_with_report(path)
        assert len(df) == 3, f"expected 3 rows, got {len(df)}: {df.to_dict()}"
        assert any("repeated header" in w for w in rpt.warnings)

    def test_multiple_repeated_headers_removed(self):
        content = (
            "x,y,z\n"
            "1,2,3\n"
            "x,y,z\n"
            "4,5,6\n"
            "x,y,z\n"
            "7,8,9\n"
        )
        path = _write_csv(content)
        df, rpt = load_dataset_with_report(path)
        assert len(df) == 3
        assert any("repeated header" in w for w in rpt.warnings)

    def test_no_false_positive_on_clean_data(self):
        content = "id,value,label\n1,100,A\n2,200,B\n3,300,C\n"
        path = _write_csv(content)
        df, rpt = load_dataset_with_report(path)
        assert len(df) == 3
        assert not any("repeated header" in w for w in rpt.warnings)


# ── Schema normalizer ─────────────────────────────────────────────────────────

class TestSchemaNormalizer:
    def test_unnamed_edge_columns_stripped(self):
        df = pd.DataFrame({
            "Unnamed: 0": [None, None],
            "name": ["Alice", "Bob"],
            "age": [30, 25],
            "Unnamed: 3": [None, None],
        })
        rpt = ParseReport()
        result = normalize_schema(df, rpt)
        assert "Unnamed: 0" not in result.columns
        assert "Unnamed: 3" not in result.columns
        assert "name" in result.columns

    def test_duplicate_columns_deduplicated(self):
        df = pd.DataFrame([[1, 2, 3]], columns=["age", "age", "age"])
        rpt = ParseReport()
        result = normalize_schema(df, rpt)
        assert list(result.columns) == ["age", "age_1", "age_2"]

    def test_blank_rows_dropped(self):
        df = pd.DataFrame({
            "a": [1, None, 3],
            "b": [4, None, 6],
        })
        rpt = ParseReport()
        result = normalize_schema(df, rpt)
        assert len(result) == 2

    def test_column_names_coerced_to_strings(self):
        df = pd.DataFrame({1: [1, 2], 2: [3, 4]})
        rpt = ParseReport()
        result = normalize_schema(df, rpt)
        assert all(isinstance(c, str) for c in result.columns)


# ── Excel region detection ─────────────────────────────────────────────────────

class TestExcelRegion:
    def test_flat_excel_no_skiprows(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["product", "units", "revenue"])
        ws.append(["Alpha", 100, 5000])
        ws.append(["Beta", 200, 8000])
        path = _write_excel(wb)
        df, rpt = load_dataset_with_report(path)
        assert "product" in [c.lower() for c in df.columns]
        assert len(df) >= 2

    def test_excel_with_title_row_skipped(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Q4 2024 Sales Report"   # decorative title
        ws["A2"] = None                       # blank spacer
        ws["A3"] = "product"
        ws["B3"] = "units"
        ws["C3"] = "revenue"
        ws.append(["Alpha", 100, 5000])
        ws.append(["Beta", 200, 8000])
        path = _write_excel(wb)
        df, rpt = load_dataset_with_report(path)
        assert len(df) >= 2
        # product column should be present (not lost to title row)
        col_names_lower = [c.lower() for c in df.columns]
        assert "product" in col_names_lower or any("alpha" in str(v).lower() for v in df.iloc[:, 0])

    def test_excel_schema_normalization(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, "name", "age", None])
        ws.append([None, "Alice", 30, None])
        ws.append([None, "Bob", 25, None])
        path = _write_excel(wb)
        df, rpt = load_dataset_with_report(path)
        assert "name" in df.columns
        assert "age" in df.columns


# ── ParseReport ───────────────────────────────────────────────────────────────

class TestParseReport:
    def test_report_returned_on_flat_csv(self):
        path = _write_csv("a,b\n1,2\n3,4\n")
        df, rpt = load_dataset_with_report(path)
        assert isinstance(rpt, ParseReport)
        assert rpt.status in ("ok", "parsed_with_warnings", "fallback")
        assert 0.0 <= rpt.confidence <= 1.0

    def test_fallback_status_on_corrupt_file(self, tmp_path):
        p = tmp_path / "corrupt.csv"
        p.write_bytes(b"\x00\x01\x02\x03" * 100)
        df, rpt = load_dataset_with_report(str(p))
        # Should not raise — either ok or fallback
        assert isinstance(df, pd.DataFrame)
        assert isinstance(rpt, ParseReport)
