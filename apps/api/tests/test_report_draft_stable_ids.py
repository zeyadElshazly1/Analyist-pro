"""
Report Draft stable insight ID tests (Task 7)
=============================================

Locks in the trust-hardening fix: ``ReportDraft.selected_insights`` selects
findings by stable ``insight_id`` rather than by fragile positional index.

These tests cover:

* ``apply_draft_to_result`` — the unit-level helper everything else routes
  through — selects by stable id, falls back to numeric indices for legacy
  drafts, and silently drops missing entries instead of selecting the wrong
  finding.
* The HTML export honours the same stable-id contract end-to-end, including
  the case where insights have been re-ordered between the draft and the
  export run.
* Excel export respects stable ids identically.
* Old drafts that recorded numeric indices keep working unchanged.
* Stable ids referencing insights that no longer exist on the run are
  dropped silently — they never coerce into a fallback wrong finding.
"""
from __future__ import annotations

import io
import json

from openpyxl import load_workbook

from app.models import AnalysisResult, ReportDraft
from app.services.reporting.draft_context import (
    _select_indices,
    apply_draft_to_result,
)
from tests.conftest import TestingSessionLocal


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_insight(idx: int, title: str, *, insight_id: str | None = None) -> dict:
    return {
        "insight_id": insight_id if insight_id is not None else f"sid_{title}",
        "title": title,
        "explanation": f"{title} — detail body line.",
        "category": "outlier",
        "severity": "high" if idx == 0 else "medium",
        "columns_used": ["salary"],
        "method_used": "IQR fence",
        "evidence": f"n={50 + idx}",
        "confidence": 0.9,
        "report_safe": True,
        "caveats": [],
        "why_it_matters": "Matters because the test says so.",
        "recommendation": f"Investigate {title}",
        "chart_suggestion": "none",
    }


def _seed_run_with_titles(
    project_id: int,
    titles: list[str],
    *,
    narrative: str = "Pipeline-generated narrative.",
) -> int:
    """Insert a report-ready AnalysisResult with deterministic insights.

    Each insight gets ``insight_id="sid_<title>"`` so tests can assert
    selection by stable id.
    """
    db = TestingSessionLocal()
    try:
        insight_results = [_make_insight(i, t) for i, t in enumerate(titles)]
        result = {
            "project_id": project_id,
            "narrative": narrative,
            "insight_results": insight_results,
            "health_score": {"total": 80, "breakdown": {"Overall": 80}},
            "dataset_summary": {"rows": 10, "columns": 3, "numeric_cols": 2, "categorical_cols": 1, "missing_pct": 0.0},
            "cleaning_report": [],
            "profile": [],
        }
        run = AnalysisResult(
            project_id=project_id,
            file_hash="seed-hash",
            result_json=json.dumps(result),
            status="report_ready",
        )
        db.add(run)
        db.commit()
        db.refresh(run)
        return run.id
    finally:
        db.close()


def _save_draft(
    client,
    project_id: int,
    headers,
    *,
    summary: str | None = None,
    selected: list | None = None,
    title: str | None = None,
):
    payload: dict = {}
    if summary is not None:
        payload["summary"] = summary
    if selected is not None:
        payload["selected_insight_ids"] = selected
    if title is not None:
        payload["title"] = title
    r = client.post(f"/reports/draft/{project_id}", json=payload, headers=headers)
    assert r.status_code == 200, r.text
    return r.json()


def _force_draft_run_link(project_id: int, run_id: int) -> None:
    db = TestingSessionLocal()
    try:
        draft = (
            db.query(ReportDraft)
            .filter(ReportDraft.project_id == project_id)
            .order_by(ReportDraft.created_at.desc())
            .first()
        )
        assert draft is not None
        draft.analysis_result_id = run_id
        db.commit()
    finally:
        db.close()


# ── _select_indices unit tests ────────────────────────────────────────────────

class TestSelectIndicesUnit:
    def test_string_ids_match_by_insight_id(self):
        source = [
            _make_insight(0, "ALPHA"),
            _make_insight(1, "BETA"),
            _make_insight(2, "GAMMA"),
        ]
        idxs = _select_indices(source, ["sid_ALPHA", "sid_GAMMA"])
        assert idxs == [0, 2]

    def test_numeric_indices_still_work_for_legacy(self):
        source = [_make_insight(0, "A"), _make_insight(1, "B")]
        idxs = _select_indices(source, [0, 1])
        assert idxs == [0, 1]

    def test_mixed_ids_and_indices(self):
        source = [
            _make_insight(0, "A"),
            _make_insight(1, "B"),
            _make_insight(2, "C"),
        ]
        # "sid_C" -> 2, 0 -> 0
        idxs = _select_indices(source, ["sid_C", 0])
        assert idxs == [2, 0]

    def test_unknown_id_is_dropped_silently(self):
        source = [_make_insight(0, "A"), _make_insight(1, "B")]
        idxs = _select_indices(source, ["sid_DOES_NOT_EXIST", "sid_A"])
        # Missing id is dropped (we'd rather show fewer findings than wrong ones).
        assert idxs == [0]

    def test_out_of_range_index_is_dropped(self):
        source = [_make_insight(0, "A")]
        idxs = _select_indices(source, [5, -1, "bad", True, False])
        # bool is filtered, ints out of range filtered, unknown id filtered.
        assert idxs == []

    def test_duplicates_are_collapsed(self):
        source = [_make_insight(0, "A"), _make_insight(1, "B")]
        idxs = _select_indices(source, ["sid_A", "sid_A", 0, 1, 1])
        assert idxs == [0, 1]

    def test_reordered_insights_select_by_id_not_position(self):
        """Critical case: same insight_id, different position → select correctly."""
        original = [
            _make_insight(0, "ALPHA"),
            _make_insight(1, "BETA"),
            _make_insight(2, "GAMMA"),
        ]
        # Re-run swaps position of GAMMA and ALPHA.
        reordered = [
            _make_insight(2, "GAMMA"),
            _make_insight(1, "BETA"),
            _make_insight(0, "ALPHA"),
        ]
        # Draft saved against original: sid_ALPHA + sid_GAMMA.
        idxs = _select_indices(reordered, ["sid_ALPHA", "sid_GAMMA"])
        # sid_ALPHA is now at index 2, sid_GAMMA at index 0 — the *findings*
        # are still ALPHA + GAMMA, not BETA + something else.
        kept = [reordered[i]["title"] for i in idxs]
        assert sorted(kept) == ["ALPHA", "GAMMA"]


# ── apply_draft_to_result with stable IDs ─────────────────────────────────────

class TestApplyDraftWithStableIds:
    def test_filters_by_stable_id(self):
        result = {
            "insight_results": [
                _make_insight(0, "A"),
                _make_insight(1, "B"),
                _make_insight(2, "C"),
            ],
        }
        applied = apply_draft_to_result(result, selected_indices=["sid_A", "sid_C"])
        titles = [i["title"] for i in applied["insight_results"]]
        assert titles == ["A", "C"]

    def test_legacy_numeric_indices_still_work(self):
        result = {
            "insight_results": [
                _make_insight(0, "A"),
                _make_insight(1, "B"),
                _make_insight(2, "C"),
            ],
        }
        applied = apply_draft_to_result(result, selected_indices=[0, 2])
        titles = [i["title"] for i in applied["insight_results"]]
        assert titles == ["A", "C"]

    def test_reordered_insights_pick_correct_findings(self):
        """The same draft against a re-ordered run still selects the same findings."""
        result_reordered = {
            "insight_results": [
                _make_insight(0, "GAMMA"),
                _make_insight(1, "BETA"),
                _make_insight(2, "ALPHA"),
            ],
        }
        # Draft was saved against an older order where ALPHA was index 0.
        applied = apply_draft_to_result(
            result_reordered,
            selected_indices=["sid_ALPHA", "sid_GAMMA"],
        )
        titles = [i["title"] for i in applied["insight_results"]]
        assert sorted(titles) == ["ALPHA", "GAMMA"]
        # Critical: BETA must NOT have been silently selected because the
        # legacy index path would have picked indices 0 and 2 instead.
        assert "BETA" not in titles

    def test_missing_stable_id_does_not_select_fallback(self):
        result = {
            "insight_results": [
                _make_insight(0, "A"),
                _make_insight(1, "B"),
            ],
        }
        applied = apply_draft_to_result(
            result,
            selected_indices=["sid_DELETED", "sid_A"],
        )
        # "sid_DELETED" is gone — exclude it; never coerce into a fallback row.
        titles = [i["title"] for i in applied["insight_results"]]
        assert titles == ["A"]

    def test_all_missing_ids_yields_empty_selection(self):
        result = {
            "insight_results": [_make_insight(0, "A")],
        }
        applied = apply_draft_to_result(
            result,
            selected_indices=["sid_GHOST_1", "sid_GHOST_2"],
        )
        # No crash, no wrong findings, just an empty filtered list.
        assert applied["insight_results"] == []

    def test_mirrors_filter_to_legacy_when_both_present(self):
        result = {
            "insight_results": [
                _make_insight(0, "A"),
                _make_insight(1, "B"),
            ],
            "insights": [
                {"finding": "A"},
                {"finding": "B"},
            ],
        }
        applied = apply_draft_to_result(result, selected_indices=["sid_B"])
        assert [i["title"] for i in applied["insight_results"]] == ["B"]
        assert [i["finding"] for i in applied["insights"]] == ["B"]

    def test_summary_still_overrides_with_stable_ids(self):
        """Task 2 behaviour must remain intact — summary edits still apply."""
        result = {
            "narrative": "auto",
            "insight_results": [_make_insight(0, "A")],
        }
        applied = apply_draft_to_result(
            result,
            draft_summary="Edited.",
            selected_indices=["sid_A"],
        )
        assert applied["narrative"] == "Edited."


# ── End-to-end: HTML export honours stable IDs across a re-run ───────────────

class TestHtmlExportWithStableIds:
    def test_stable_id_selection_survives_reordered_rerun(
        self, client, uploaded_project, consultant_auth_headers
    ):
        """Save a draft against one run, swap insight order on the next run,
        and confirm the export still includes the originally selected
        findings — never the wrong ones."""
        pid = uploaded_project["id"]

        # Run #1: ALPHA, BETA, GAMMA in that order.
        run1 = _seed_run_with_titles(pid, ["ALPHA", "BETA", "GAMMA"])

        # Save draft selecting ALPHA + GAMMA by stable id.
        _save_draft(
            client, pid, consultant_auth_headers,
            summary="EDITED_SUMMARY_TOKEN — keep ALPHA and GAMMA only.",
            selected=["sid_ALPHA", "sid_GAMMA"],
        )
        _force_draft_run_link(pid, run1)

        # Run #2 reorders insights — same set, different positions.
        # The draft is still pinned to run1, so this also exercises the
        # "pinned run" path from Task 2.
        _seed_run_with_titles(pid, ["GAMMA", "BETA", "ALPHA"])

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        body = r.text

        assert "EDITED_SUMMARY_TOKEN" in body
        assert "ALPHA" in body
        assert "GAMMA" in body
        # BETA must never appear — it would only show up if the old
        # index-based selection silently included index 1.
        assert ">BETA<" not in body and " BETA " not in body, (
            "BETA was deselected — it must not appear in HTML even after a re-run "
            "that re-orders the insight list."
        )

    def test_legacy_numeric_indices_still_export_correctly(
        self, client, uploaded_project, consultant_auth_headers
    ):
        """A pre-existing draft that saved numeric indices must keep working."""
        pid = uploaded_project["id"]
        _seed_run_with_titles(pid, ["LEGACY_KEEP", "LEGACY_DROP", "LEGACY_KEEP_2"])

        _save_draft(
            client, pid, consultant_auth_headers,
            summary="Legacy summary token.",
            selected=[0, 2],  # numeric indices — legacy shape
        )

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200
        body = r.text
        assert "LEGACY_KEEP" in body
        assert "LEGACY_KEEP_2" in body
        assert "LEGACY_DROP" not in body

    def test_unknown_stable_ids_do_not_crash_export(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        _seed_run_with_titles(pid, ["LIVE_FINDING"])

        _save_draft(
            client, pid, consultant_auth_headers,
            summary="Summary still exports even when selection points to deleted findings.",
            selected=["sid_NOT_THERE", "sid_ALSO_GONE"],
        )

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        body = r.text
        # Edited summary still appears (Task 2 contract preserved).
        assert "Summary still exports" in body
        # Live finding is not silently selected as a fallback.
        assert "LIVE_FINDING" not in body


# ── End-to-end: Excel export honours stable IDs ──────────────────────────────

class TestExcelExportWithStableIds:
    def test_excel_filters_by_stable_id(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        _seed_run_with_titles(pid, [
            "EXCEL_KEEP_A",
            "EXCEL_DROP_B",
            "EXCEL_KEEP_C",
        ])

        _save_draft(
            client, pid, consultant_auth_headers,
            summary="EXCEL_STABLE_ID_TOKEN.",
            selected=["sid_EXCEL_KEEP_A", "sid_EXCEL_KEEP_C"],
        )

        r = client.get(f"/reports/export/{pid}?format=xlsx", headers=consultant_auth_headers)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content), data_only=True)

        summary_text = "\n".join(
            str(c.value) for row in wb["Summary"].iter_rows() for c in row if c.value
        )
        assert "EXCEL_STABLE_ID_TOKEN" in summary_text

        insights_text = "\n".join(
            str(c.value) for row in wb["Insights"].iter_rows() for c in row if c.value
        )
        assert "EXCEL_KEEP_A" in insights_text
        assert "EXCEL_KEEP_C" in insights_text
        assert "EXCEL_DROP_B" not in insights_text


# ── Draft response round-trip ────────────────────────────────────────────────

class TestDraftRoundTrip:
    def test_draft_returns_stable_ids_when_saved_with_stable_ids(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        _seed_run_with_titles(pid, ["RT_A", "RT_B", "RT_C"])

        body = _save_draft(
            client, pid, consultant_auth_headers,
            selected=["sid_RT_A", "sid_RT_C"],
        )
        # Stored shape preserved.
        assert body["selected_insight_ids"] == ["sid_RT_A", "sid_RT_C"]

        # Embedded report_result.included_insights resolved correctly.
        included = body.get("report_result", {}).get("included_insights", [])
        titles = [i["title"] for i in included]
        assert sorted(titles) == ["RT_A", "RT_C"]

    def test_draft_returns_numeric_indices_when_saved_legacy(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        _seed_run_with_titles(pid, ["RT_A", "RT_B", "RT_C"])

        body = _save_draft(
            client, pid, consultant_auth_headers,
            selected=[0, 2],
        )
        assert body["selected_insight_ids"] == [0, 2]

        included = body.get("report_result", {}).get("included_insights", [])
        titles = [i["title"] for i in included]
        assert sorted(titles) == ["RT_A", "RT_C"]
