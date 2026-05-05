"""
Report Draft Export tests
=========================

Verifies that the saved Report Builder draft drives what comes out of the
``/reports/export`` and ``/reports/preview`` endpoints:

- Edited summary appears in HTML and Excel exports.
- Selected findings appear; deselected findings do not.
- Preview uses the same (filtered) context as export.
- A draft pinned to an older analysis run does not silently switch to the
  latest run.
- A draft whose ``analysis_result_id`` points at a different project's run
  is rejected and falls back to the project's own latest run.
- The draft-applied helper in isolation behaves correctly.
- build_context exposes selected_chart_payloads resolved from stored chart
  blocks.

These tests are the contract for the launch-hardening report-export fix.
"""
from __future__ import annotations

import io
import json

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.models import AnalysisResult, ReportDraft
from app.services.reporting.charts import resolve_selected_chart_payloads
from app.services.reporting.context import build_context
from app.services.reporting.draft_context import apply_draft_to_result
from tests.conftest import TestingSessionLocal


# ── Helpers ───────────────────────────────────────────────────────────────────

def _seed_run(
    project_id: int,
    *,
    insight_titles: list[str],
    narrative: str = "Pipeline-generated narrative.",
) -> int:
    """Insert an AnalysisResult directly with deterministic insight_results.

    Returns the run id.  Each insight has stable, easy-to-grep ``title`` and
    ``explanation`` strings so tests can assert presence/absence.
    """
    db = TestingSessionLocal()
    try:
        insight_results = []
        for idx, title in enumerate(insight_titles):
            insight_results.append({
                "insight_id": f"ins_{idx}",
                "title": title,
                "explanation": f"{title} — detail body line.",
                "category": "outlier",
                "severity": "high" if idx == 0 else "medium",
                "columns_used": ["salary"],
                "method_used": "IQR fence",
                "evidence": f"n={50 + idx}",
                "confidence": 0.9,
                "report_safe": True,
                "caveats": [],
                "why_it_matters": "Matters because the test says so.",
                "recommendation": f"Investigate {title}",
                "chart_suggestion": "none",
            })

        result = {
            "project_id": project_id,
            "narrative": narrative,
            "insight_results": insight_results,
            "health_score": {"total": 80, "breakdown": {"Overall": 80}},
            "dataset_summary": {"rows": 10, "columns": 3, "numeric_cols": 2, "categorical_cols": 1, "missing_pct": 0.0},
            "cleaning_report": [],
            "profile": [],
        }
        run = AnalysisResult(
            project_id=project_id,
            file_hash="seed-hash",
            result_json=json.dumps(result),
            status="report_ready",
        )
        db.add(run)
        db.commit()
        db.refresh(run)
        return run.id
    finally:
        db.close()


def _save_draft(
    client,
    project_id: int,
    headers,
    *,
    summary: str | None = None,
    selected: list[int] | None = None,
    title: str | None = None,
):
    """POST to /reports/draft/{pid}; returns the response body."""
    payload: dict = {}
    if summary is not None:
        payload["summary"] = summary
    if selected is not None:
        payload["selected_insight_ids"] = selected
    if title is not None:
        payload["title"] = title
    r = client.post(f"/reports/draft/{project_id}", json=payload, headers=headers)
    assert r.status_code == 200, r.text
    return r.json()


def _link_draft_to_run(project_id: int, run_id: int) -> None:
    """Force the latest draft for ``project_id`` to point at ``run_id``."""
    db = TestingSessionLocal()
    try:
        draft = (
            db.query(ReportDraft)
            .filter(ReportDraft.project_id == project_id)
            .order_by(ReportDraft.created_at.desc())
            .first()
        )
        assert draft is not None, "No draft exists yet — call _save_draft first"
        draft.analysis_result_id = run_id
        db.commit()
    finally:
        db.close()


# ── apply_draft_to_result unit tests ──────────────────────────────────────────

class TestApplyDraftToResultUnit:
    def test_filters_insight_results_by_selected_indices(self):
        result = {
            "narrative": "original",
            "insight_results": [
                {"title": "Finding A"},
                {"title": "Finding B"},
                {"title": "Finding C"},
            ],
        }
        applied = apply_draft_to_result(result, selected_indices=[0, 2])
        titles = [i["title"] for i in applied["insight_results"]]
        assert titles == ["Finding A", "Finding C"]
        # Original input must not be mutated.
        assert len(result["insight_results"]) == 3

    def test_filters_legacy_insights_when_canonical_missing(self):
        result = {"insights": [{"finding": "X"}, {"finding": "Y"}]}
        applied = apply_draft_to_result(result, selected_indices=[1])
        assert [i["finding"] for i in applied["insights"]] == ["Y"]

    def test_mirrors_filter_to_legacy_when_both_present(self):
        result = {
            "insight_results": [{"title": "A"}, {"title": "B"}, {"title": "C"}],
            "insights":         [{"finding": "A"}, {"finding": "B"}, {"finding": "C"}],
        }
        applied = apply_draft_to_result(result, selected_indices=[1])
        assert [i["title"]   for i in applied["insight_results"]] == ["B"]
        assert [i["finding"] for i in applied["insights"]]        == ["B"]

    def test_replaces_narrative_with_draft_summary(self):
        result = {"narrative": "auto", "insight_results": []}
        applied = apply_draft_to_result(result, draft_summary="My edited summary")
        assert applied["narrative"] == "My edited summary"
        assert applied["draft_summary"] == "My edited summary"

    def test_blank_summary_does_not_overwrite_narrative(self):
        result = {"narrative": "auto", "insight_results": []}
        applied = apply_draft_to_result(result, draft_summary="   ")
        assert applied["narrative"] == "auto"
        assert "draft_summary" not in applied

    def test_none_selected_indices_keeps_full_list(self):
        result = {"insight_results": [{"title": "A"}, {"title": "B"}]}
        applied = apply_draft_to_result(result, selected_indices=None)
        assert len(applied["insight_results"]) == 2

    def test_invalid_indices_are_clamped(self):
        result = {"insight_results": [{"title": "A"}, {"title": "B"}]}
        applied = apply_draft_to_result(result, selected_indices=[0, 99, -1, "bad"])  # type: ignore[list-item]
        assert [i["title"] for i in applied["insight_results"]] == ["A"]


# ── End-to-end: HTML export honours the draft ────────────────────────────────

class TestHtmlExportUsesDraft:
    def test_edited_summary_and_selected_findings_appear(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        _seed_run(pid, insight_titles=[
            "ALPHA_KEEP_FINDING",
            "BETA_DESELECTED_FINDING",
            "GAMMA_KEEP_FINDING",
        ])

        edited = "EDITED_SUMMARY_TOKEN — this is what the consultant wrote."
        _save_draft(
            client, pid, consultant_auth_headers,
            summary=edited,
            selected=[0, 2],  # keep ALPHA + GAMMA, drop BETA
        )

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        body = r.text

        assert edited in body, "Edited summary must appear in HTML export"
        assert "ALPHA_KEEP_FINDING" in body
        assert "GAMMA_KEEP_FINDING" in body
        assert "BETA_DESELECTED_FINDING" not in body, (
            "Deselected finding must not appear in HTML export"
        )

    def test_html_uses_pinned_run_not_latest(
        self, client, uploaded_project, consultant_auth_headers
    ):
        """A draft pinned to an older run must not export the latest run's data."""
        pid = uploaded_project["id"]

        # Older run — what the draft was built against.
        old_run_id = _seed_run(pid, insight_titles=["OLD_RUN_FINDING"])

        # Save a draft against the old run (currently the "latest").
        _save_draft(
            client, pid, consultant_auth_headers,
            summary="Summary tied to OLD run.",
            selected=[0],
        )
        _link_draft_to_run(pid, old_run_id)

        # Now seed a *newer* run with a clearly different finding.
        _seed_run(pid, insight_titles=["NEW_RUN_FINDING"])

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        assert "OLD_RUN_FINDING" in r.text
        assert "NEW_RUN_FINDING" not in r.text, (
            "Export must stay on the draft-linked analysis run, not silently "
            "switch to the latest run."
        )


# ── End-to-end: Excel export honours the draft ───────────────────────────────

class TestExcelExportUsesDraft:
    def test_edited_summary_and_selected_findings_appear(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        _seed_run(pid, insight_titles=[
            "EXCEL_ALPHA_KEEP",
            "EXCEL_BETA_DROP",
            "EXCEL_GAMMA_KEEP",
        ])

        edited = "EXCEL_SUMMARY_EDIT_TOKEN — written by the consultant."
        _save_draft(
            client, pid, consultant_auth_headers,
            summary=edited,
            selected=[0, 2],
        )

        r = client.get(f"/reports/export/{pid}?format=xlsx", headers=consultant_auth_headers)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content), data_only=True)

        # Summary sheet should contain the edited executive summary.
        summary_text = "\n".join(
            str(c.value) for row in wb["Summary"].iter_rows() for c in row if c.value
        )
        assert edited in summary_text, "Edited summary must appear on the Summary sheet"

        # Insights sheet should include only selected findings.
        insights_text = "\n".join(
            str(c.value) for row in wb["Insights"].iter_rows() for c in row if c.value
        )
        assert "EXCEL_ALPHA_KEEP" in insights_text
        assert "EXCEL_GAMMA_KEEP" in insights_text
        assert "EXCEL_BETA_DROP" not in insights_text


# ── Preview parity ────────────────────────────────────────────────────────────

class TestPreviewMatchesExport:
    def test_preview_returns_filtered_context(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        _seed_run(pid, insight_titles=[
            "PREVIEW_ALPHA",
            "PREVIEW_BETA_DROP",
            "PREVIEW_GAMMA",
        ])

        edited = "PREVIEW_SUMMARY_TOKEN."
        _save_draft(
            client, pid, consultant_auth_headers,
            summary=edited,
            selected=[0, 2],
        )

        r = client.get(f"/reports/preview/{pid}", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        body = r.json()

        assert body["narrative"] == edited
        titles = [i.get("title") for i in body.get("insight_results", [])]
        assert titles == ["PREVIEW_ALPHA", "PREVIEW_GAMMA"]


# ── Cross-project draft / analysis_result mismatch is rejected ───────────────

class TestCrossProjectMismatchRejected:
    def test_draft_with_other_project_run_falls_back_to_own_latest(
        self, client, uploaded_project, consultant_auth_headers, csv_bytes, auth_headers
    ):
        """A draft whose analysis_result_id points at another project's run
        must not pull data from that other project — the resolver falls back
        to this project's own latest run.
        """
        pid_a = uploaded_project["id"]

        # Project B for the same user — different project, different run.
        r = client.post("/projects", json={"name": "Project B"}, headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        pid_b = r.json()["id"]
        client.post(
            "/upload",
            files={"file": ("data.csv", io.BytesIO(csv_bytes), "text/csv")},
            data={"project_id": str(pid_b)},
            headers=consultant_auth_headers,
        )
        run_b_id = _seed_run(pid_b, insight_titles=["PROJECT_B_FINDING"])

        # Project A: own analysis exists.
        run_a_id = _seed_run(pid_a, insight_titles=["PROJECT_A_FINDING"])

        # Save a draft for project A and forcibly point it at project B's run id.
        _save_draft(
            client, pid_a, consultant_auth_headers,
            summary="Cross-project summary.",
            selected=[0],
        )
        _link_draft_to_run(pid_a, run_b_id)

        r = client.get(f"/reports/export/{pid_a}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        body = r.text
        assert "PROJECT_A_FINDING" in body
        assert "PROJECT_B_FINDING" not in body, (
            "A tampered draft must not be able to pull data from another "
            "project's analysis run."
        )


# ── Backwards compatibility: no draft yet ─────────────────────────────────────

class TestNoDraftFallback:
    def test_export_without_draft_uses_latest_run(
        self, client, uploaded_project, consultant_auth_headers
    ):
        """When no draft has been saved, export still works against the
        latest run — preserves existing behaviour for users who never opened
        the Report Builder."""
        pid = uploaded_project["id"]
        _seed_run(pid, insight_titles=["DEFAULT_FINDING_NO_DRAFT"])

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        assert "DEFAULT_FINDING_NO_DRAFT" in r.text


# ── apply_draft_to_result: selected_chart_ids propagation ─────────────────────

class TestApplyDraftSelectedChartIds:
    def test_stores_chart_ids_in_result(self):
        result = {"insight_results": []}
        applied = apply_draft_to_result(result, selected_chart_ids=["c1", "c2"])
        assert applied["selected_chart_ids"] == ["c1", "c2"]

    def test_none_chart_ids_does_not_add_key(self):
        result = {"insight_results": []}
        applied = apply_draft_to_result(result, selected_chart_ids=None)
        assert "selected_chart_ids" not in applied

    def test_empty_list_stored_explicitly(self):
        result = {"insight_results": []}
        applied = apply_draft_to_result(result, selected_chart_ids=[])
        assert applied["selected_chart_ids"] == []

    def test_does_not_mutate_original(self):
        result = {"insight_results": [], "charts": [{"chart_id": "x"}]}
        apply_draft_to_result(result, selected_chart_ids=["x"])
        assert "selected_chart_ids" not in result

    def test_combined_with_insight_selection(self):
        result = {
            "insight_results": [
                {"insight_id": "a", "title": "A"},
                {"insight_id": "b", "title": "B"},
            ],
        }
        applied = apply_draft_to_result(
            result, selected_indices=["a"], selected_chart_ids=["c9"]
        )
        assert [i["title"] for i in applied["insight_results"]] == ["A"]
        assert applied["selected_chart_ids"] == ["c9"]


# ── resolve_selected_chart_payloads unit tests ────────────────────────────────

class TestResolveSelectedChartPayloads:
    def test_returns_empty_when_no_key(self):
        assert resolve_selected_chart_payloads({}) == []

    def test_returns_empty_when_key_is_empty_list(self):
        assert resolve_selected_chart_payloads({"selected_chart_ids": []}) == []

    def test_resolves_string_id_from_charts_block(self):
        result = {
            "selected_chart_ids": ["c1"],
            "charts": [{"chart_id": "c1", "title": "Revenue", "type": "bar"}],
        }
        out = resolve_selected_chart_payloads(result)
        assert len(out) == 1
        assert out[0]["chart_id"] == "c1"
        assert out[0]["title"] == "Revenue"
        assert out[0]["chart_type"] == "bar"

    def test_resolves_from_chart_results_block(self):
        result = {
            "selected_chart_ids": ["cr1"],
            "chart_results": [{"chart_id": "cr1", "title": "Sector split", "type": "pie"}],
        }
        out = resolve_selected_chart_payloads(result)
        assert out[0]["chart_id"] == "cr1"
        assert out[0]["chart_type"] == "pie"

    def test_resolves_from_suggested_charts_block(self):
        result = {
            "selected_chart_ids": ["s1"],
            "suggested_charts": [{"chart_id": "s1", "title": "Trend", "type": "line"}],
        }
        out = resolve_selected_chart_payloads(result)
        assert out[0]["chart_id"] == "s1"

    def test_resolves_from_chart_gallery_block(self):
        result = {
            "selected_chart_ids": ["g1"],
            "chart_gallery": [{"chart_id": "g1", "title": "Gallery item", "type": "scatter"}],
        }
        out = resolve_selected_chart_payloads(result)
        assert out[0]["chart_id"] == "g1"

    def test_skips_unresolvable_ids(self):
        result = {
            "selected_chart_ids": ["known", "ghost"],
            "charts": [{"chart_id": "known", "title": "Known chart", "type": "bar"}],
        }
        out = resolve_selected_chart_payloads(result)
        assert len(out) == 1
        assert out[0]["chart_id"] == "known"

    def test_deduplicates_by_chart_id(self):
        result = {
            "selected_chart_ids": ["dup", "dup"],
            "charts": [{"chart_id": "dup", "title": "Dup", "type": "bar"}],
        }
        out = resolve_selected_chart_payloads(result)
        assert len(out) == 1

    def test_preserves_order_matching_selected_ids(self):
        result = {
            "selected_chart_ids": ["b", "a"],
            "charts": [
                {"chart_id": "a", "title": "A"},
                {"chart_id": "b", "title": "B"},
            ],
        }
        out = resolve_selected_chart_payloads(result)
        assert [e["chart_id"] for e in out] == ["b", "a"]

    def test_legacy_integer_index_resolved(self):
        result = {
            "selected_chart_ids": [1],
            "charts": [
                {"title": "First", "type": "bar"},
                {"title": "Second", "type": "line"},
            ],
        }
        out = resolve_selected_chart_payloads(result)
        assert len(out) == 1
        assert out[0]["chart_id"] == "idx_1"
        assert out[0]["title"] == "Second"
        assert out[0]["chart_type"] == "line"

    def test_integer_index_out_of_range_skipped(self):
        result = {
            "selected_chart_ids": [99],
            "charts": [{"chart_id": "only", "title": "Only", "type": "bar"}],
        }
        assert resolve_selected_chart_payloads(result) == []

    def test_integer_index_resolves_to_existing_chart_id(self):
        result = {
            "selected_chart_ids": [0],
            "charts": [{"chart_id": "has_id", "title": "Has ID", "type": "bar"}],
        }
        out = resolve_selected_chart_payloads(result)
        assert out[0]["chart_id"] == "has_id"

    def test_dict_slot_resolved_by_chart_id_field(self):
        result = {
            "selected_chart_ids": [{"chart_id": "d1", "title": "override"}],
            "charts": [{"chart_id": "d1", "title": "Stored", "type": "bar"}],
        }
        out = resolve_selected_chart_payloads(result)
        assert out[0]["chart_id"] == "d1"

    def test_preserves_all_stored_fields(self):
        result = {
            "selected_chart_ids": ["rich"],
            "charts": [{
                "chart_id": "rich",
                "title": "Rich chart",
                "type": "bar",
                "data": {"labels": ["A"], "datasets": [{"data": [1]}]},
                "options": {"responsive": True},
            }],
        }
        out = resolve_selected_chart_payloads(result)
        assert "data" in out[0]
        assert "options" in out[0]

    def test_normalises_chart_type_from_type_field(self):
        result = {
            "selected_chart_ids": ["t1"],
            "charts": [{"chart_id": "t1", "title": "T", "type": "scatter"}],
        }
        out = resolve_selected_chart_payloads(result)
        assert out[0]["chart_type"] == "scatter"

    def test_unknown_chart_type_when_missing(self):
        result = {
            "selected_chart_ids": ["no_type"],
            "charts": [{"chart_id": "no_type", "title": "No type chart"}],
        }
        out = resolve_selected_chart_payloads(result)
        assert out[0]["chart_type"] == "unknown"

    def test_booleans_in_selected_ids_are_skipped(self):
        result = {
            "selected_chart_ids": [True, False, "real"],
            "charts": [{"chart_id": "real", "title": "R", "type": "bar"}],
        }
        out = resolve_selected_chart_payloads(result)
        assert len(out) == 1
        assert out[0]["chart_id"] == "real"

    def test_empty_charts_block_falls_through_to_chart_results(self):
        result = {
            "selected_chart_ids": ["x"],
            "charts": [],
            "chart_results": [{"chart_id": "x", "title": "X", "type": "bar"}],
        }
        out = resolve_selected_chart_payloads(result)
        assert out[0]["chart_id"] == "x"


# ── build_context: selected_chart_payloads integration ───────────────────────

_MINIMAL_DF = pd.DataFrame({"a": [1, 2, 3]})


class TestBuildContextSelectedChartPayloads:
    def test_key_present_and_empty_when_no_selection(self):
        result = {
            "insight_results": [],
            "health_score": {"total": 80, "breakdown": {"Overall": 80}},
        }
        ctx = build_context(_MINIMAL_DF, result, "Test")
        assert "selected_chart_payloads" in ctx
        assert ctx["selected_chart_payloads"] == []

    def test_resolves_payloads_when_selection_present(self):
        result = {
            "selected_chart_ids": ["c1"],
            "charts": [{"chart_id": "c1", "title": "Revenue", "type": "bar"}],
            "insight_results": [],
            "health_score": {"total": 70},
        }
        ctx = build_context(_MINIMAL_DF, result, "Test")
        payloads = ctx["selected_chart_payloads"]
        assert len(payloads) == 1
        assert payloads[0]["chart_id"] == "c1"
        assert payloads[0]["title"] == "Revenue"
        assert payloads[0]["chart_type"] == "bar"

    def test_skips_unresolvable_ids_in_context(self):
        result = {
            "selected_chart_ids": ["present", "absent"],
            "charts": [{"chart_id": "present", "title": "Here", "type": "line"}],
            "insight_results": [],
            "health_score": {"total": 70},
        }
        ctx = build_context(_MINIMAL_DF, result, "Test")
        ids = [p["chart_id"] for p in ctx["selected_chart_payloads"]]
        assert ids == ["present"]

    def test_legacy_health_and_missing_charts_unchanged(self):
        result = {
            "health_score": {"total": 65, "breakdown": {"Completeness": 65}},
            "profile": [{"name": "col_a", "missing_pct": 20}],
            "insight_results": [],
        }
        ctx = build_context(_MINIMAL_DF, result, "Test")
        assert ctx["health_chart"] is not None
        assert ctx["missing_chart"] is not None
        assert "selected_chart_payloads" in ctx

    def test_full_pipeline_apply_draft_then_build_context(self):
        analysis_result = {
            "narrative": "original",
            "insight_results": [
                {"insight_id": "ins_a", "title": "Insight A", "severity": "high"},
            ],
            "charts": [
                {"chart_id": "ch_1", "title": "Top performers", "type": "bar"},
                {"chart_id": "ch_2", "title": "Risk scatter", "type": "scatter"},
            ],
            "health_score": {"total": 75},
        }
        applied = apply_draft_to_result(
            analysis_result,
            draft_summary="Consultant summary.",
            selected_indices=["ins_a"],
            selected_chart_ids=["ch_2", "ch_1"],
        )
        ctx = build_context(_MINIMAL_DF, applied, "Project X")
        payloads = ctx["selected_chart_payloads"]
        assert [p["chart_id"] for p in payloads] == ["ch_2", "ch_1"]
        assert payloads[0]["chart_type"] == "scatter"
        assert payloads[1]["chart_type"] == "bar"
        assert ctx["narrative"] == "Consultant summary."


# ── HTML export: Chart Gallery section ───────────────────────────────────────

class TestHtmlExportChartGallery:
    def _seed_run_with_charts(self, project_id: int) -> int:
        db = TestingSessionLocal()
        try:
            result = {
                "project_id": project_id,
                "narrative": "Narrative text.",
                "insight_results": [
                    {
                        "insight_id": "ins_0",
                        "title": "Finding A",
                        "severity": "high",
                        "explanation": "Detail.",
                        "recommendation": "Fix it",
                    }
                ],
                "health_score": {"total": 80, "breakdown": {"Overall": 80}},
                "dataset_summary": {"rows": 10, "columns": 3, "numeric_cols": 2, "categorical_cols": 1, "missing_pct": 0.0},
                "cleaning_report": [],
                "profile": [],
                "charts": [
                    {
                        "chart_id": "gallery_a",
                        "title": "CHART_GALLERY_A_TITLE",
                        "type": "bar",
                        "chart_type": "bar",
                        "data": {"labels": ["X", "Y"], "datasets": [{"label": "s", "data": [1, 2]}]},
                    },
                    {
                        "chart_id": "gallery_b",
                        "title": "CHART_GALLERY_B_TITLE",
                        "type": "scatter",
                        "chart_type": "scatter",
                        "data": {"labels": ["P", "Q"], "datasets": [{"label": "t", "data": [3, 4]}]},
                    },
                ],
            }
            run = AnalysisResult(
                project_id=project_id,
                file_hash="chart-seed-hash",
                result_json=json.dumps(result),
                status="report_ready",
            )
            db.add(run)
            db.commit()
            db.refresh(run)
            return run.id
        finally:
            db.close()

    def _save_draft_with_charts(self, client, project_id: int, headers, *, chart_ids: list) -> None:
        _save_draft(client, project_id, headers, summary="Gallery summary.")
        db = TestingSessionLocal()
        try:
            draft = (
                db.query(ReportDraft)
                .filter(ReportDraft.project_id == project_id)
                .order_by(ReportDraft.created_at.desc())
                .first()
            )
            assert draft is not None
            draft.selected_chart_ids_json = json.dumps(chart_ids)
            db.commit()
        finally:
            db.close()

    def test_chart_gallery_section_present_when_charts_selected(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        self._save_draft_with_charts(client, pid, consultant_auth_headers, chart_ids=["gallery_a"])

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        assert "Chart Gallery" in r.text

    def test_chart_titles_rendered_in_gallery(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        self._save_draft_with_charts(
            client, pid, consultant_auth_headers, chart_ids=["gallery_a", "gallery_b"]
        )

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        assert "CHART_GALLERY_A_TITLE" in r.text
        assert "CHART_GALLERY_B_TITLE" in r.text

    def test_chart_type_badge_rendered_uppercased(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        self._save_draft_with_charts(client, pid, consultant_auth_headers, chart_ids=["gallery_a"])

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        assert "BAR" in r.text

    def test_charts_rendered_in_selection_order(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        # Select B before A to verify ordering respects selected_chart_ids
        self._save_draft_with_charts(
            client, pid, consultant_auth_headers, chart_ids=["gallery_b", "gallery_a"]
        )

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        pos_b = r.text.index("CHART_GALLERY_B_TITLE")
        pos_a = r.text.index("CHART_GALLERY_A_TITLE")
        assert pos_b < pos_a, "Gallery B must appear before Gallery A (selection order)"

    def test_unselected_chart_absent_from_gallery(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        # Only select gallery_a
        self._save_draft_with_charts(client, pid, consultant_auth_headers, chart_ids=["gallery_a"])

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        assert "CHART_GALLERY_A_TITLE" in r.text
        assert "CHART_GALLERY_B_TITLE" not in r.text

    def test_no_gallery_section_when_no_charts_selected(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        self._save_draft_with_charts(client, pid, consultant_auth_headers, chart_ids=[])

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        assert "Chart Gallery" not in r.text


# ── Excel export: Selected Charts manifest sheet ─────────────────────────────

class TestExcelChartManifestSheet:
    """Verify the 'Selected Charts' worksheet is produced correctly."""

    # Re-use the same chart-seeding helper as TestHtmlExportChartGallery.
    def _seed_run_with_charts(self, project_id: int) -> int:
        db = TestingSessionLocal()
        try:
            result = {
                "project_id": project_id,
                "narrative": "Narrative text.",
                "insight_results": [
                    {
                        "insight_id": "ins_0",
                        "title": "Finding A",
                        "severity": "high",
                        "explanation": "Detail.",
                        "recommendation": "Fix it",
                    }
                ],
                "health_score": {"total": 80, "breakdown": {"Overall": 80}},
                "dataset_summary": {"rows": 10, "columns": 3, "numeric_cols": 2, "categorical_cols": 1, "missing_pct": 0.0},
                "cleaning_report": [],
                "profile": [],
                "charts": [
                    {
                        "chart_id": "xls_a",
                        "title": "XLS_CHART_A_TITLE",
                        "type": "bar",
                        "chart_type": "bar",
                        "data": {"labels": ["X", "Y"], "datasets": [{"label": "s", "data": [1, 2]}]},
                    },
                    {
                        "chart_id": "xls_b",
                        "title": "XLS_CHART_B_TITLE",
                        "type": "line",
                        "chart_type": "line",
                        "data": {"labels": ["P", "Q"], "datasets": [{"label": "t", "data": [3, 4]}]},
                    },
                    {
                        "chart_id": "xls_c",
                        "title": "XLS_CHART_C_NODATA",
                        "type": "scatter",
                        "chart_type": "scatter",
                    },
                ],
            }
            run = AnalysisResult(
                project_id=project_id,
                file_hash="xls-chart-seed-hash",
                result_json=json.dumps(result),
                status="report_ready",
            )
            db.add(run)
            db.commit()
            db.refresh(run)
            return run.id
        finally:
            db.close()

    def _save_draft_with_charts(self, client, project_id: int, headers, *, chart_ids: list) -> None:
        _save_draft(client, project_id, headers, summary="Excel gallery summary.")
        db = TestingSessionLocal()
        try:
            draft = (
                db.query(ReportDraft)
                .filter(ReportDraft.project_id == project_id)
                .order_by(ReportDraft.created_at.desc())
                .first()
            )
            assert draft is not None
            draft.selected_chart_ids_json = json.dumps(chart_ids)
            db.commit()
        finally:
            db.close()

    def _get_xlsx(self, client, project_id: int, headers):
        r = client.get(f"/reports/export/{project_id}?format=xlsx", headers=headers)
        assert r.status_code == 200, r.text
        return load_workbook(io.BytesIO(r.content), data_only=True)

    def test_selected_charts_sheet_present_when_charts_selected(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        self._save_draft_with_charts(client, pid, consultant_auth_headers, chart_ids=["xls_a"])

        wb = self._get_xlsx(client, pid, consultant_auth_headers)
        assert "Selected Charts" in wb.sheetnames

    def test_sheet_includes_selected_chart_titles(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        self._save_draft_with_charts(
            client, pid, consultant_auth_headers, chart_ids=["xls_a", "xls_b"]
        )

        wb = self._get_xlsx(client, pid, consultant_auth_headers)
        text = "\n".join(
            str(c.value) for row in wb["Selected Charts"].iter_rows() for c in row if c.value
        )
        assert "XLS_CHART_A_TITLE" in text
        assert "XLS_CHART_B_TITLE" in text

    def test_sheet_preserves_selection_order(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        # Select B before A
        self._save_draft_with_charts(
            client, pid, consultant_auth_headers, chart_ids=["xls_b", "xls_a"]
        )

        wb = self._get_xlsx(client, pid, consultant_auth_headers)
        ws = wb["Selected Charts"]
        # Row 1 is the header; data starts at row 2.
        titles = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
        titles = [t for t in titles if t]
        assert titles.index("XLS_CHART_B_TITLE") < titles.index("XLS_CHART_A_TITLE")

    def test_sheet_excludes_unselected_charts(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        # Only xls_a selected
        self._save_draft_with_charts(client, pid, consultant_auth_headers, chart_ids=["xls_a"])

        wb = self._get_xlsx(client, pid, consultant_auth_headers)
        text = "\n".join(
            str(c.value) for row in wb["Selected Charts"].iter_rows() for c in row if c.value
        )
        assert "XLS_CHART_A_TITLE" in text
        assert "XLS_CHART_B_TITLE" not in text

    def test_no_selected_charts_sheet_when_empty_selection(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        self._save_draft_with_charts(client, pid, consultant_auth_headers, chart_ids=[])

        wb = self._get_xlsx(client, pid, consultant_auth_headers)
        assert "Selected Charts" not in wb.sheetnames

    def test_data_status_available_and_unavailable(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        # xls_a has data; xls_c has no data block
        self._save_draft_with_charts(
            client, pid, consultant_auth_headers, chart_ids=["xls_a", "xls_c"]
        )

        wb = self._get_xlsx(client, pid, consultant_auth_headers)
        ws = wb["Selected Charts"]
        statuses = [ws.cell(row=r, column=5).value for r in range(2, ws.max_row + 1)]
        statuses = [s for s in statuses if s]
        assert "available" in statuses
        assert "unavailable" in statuses

    def test_existing_sheets_unchanged(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        _seed_run(pid, insight_titles=["EXISTING_FINDING"])
        self._seed_run_with_charts(pid)
        self._save_draft_with_charts(client, pid, consultant_auth_headers, chart_ids=["xls_a"])

        wb = self._get_xlsx(client, pid, consultant_auth_headers)
        for expected in ("Summary", "Insights", "Column Profiles", "Cleaning Report"):
            assert expected in wb.sheetnames, f"Sheet '{expected}' must still exist"


# ── HTML: PDF-safe fallback markup for selected charts ───────────────────────

class TestPdfFallbackMarkup:
    """Verify the chart-pdf-fallback block is present in HTML export."""

    def _seed_run_with_charts(self, project_id: int) -> int:
        db = TestingSessionLocal()
        try:
            result = {
                "project_id": project_id,
                "narrative": "Narrative.",
                "insight_results": [],
                "health_score": {"total": 80, "breakdown": {"Overall": 80}},
                "dataset_summary": {"rows": 5, "columns": 2, "numeric_cols": 1, "categorical_cols": 1, "missing_pct": 0.0},
                "cleaning_report": [],
                "profile": [],
                "charts": [
                    {
                        "chart_id": "pdf_chart_a",
                        "title": "PDF_CHART_A_TITLE",
                        "type": "bar",
                        "chart_type": "bar",
                        "data": {"labels": ["X"], "datasets": [{"data": [1]}]},
                    },
                    {
                        "chart_id": "pdf_chart_b",
                        "title": "PDF_CHART_B_TITLE",
                        "type": "line",
                        "chart_type": "line",
                        "data": {"labels": ["Y"], "datasets": [{"data": [2]}]},
                    },
                ],
            }
            run = AnalysisResult(
                project_id=project_id,
                file_hash="pdf-seed-hash",
                result_json=json.dumps(result),
                status="report_ready",
            )
            db.add(run)
            db.commit()
            db.refresh(run)
            return run.id
        finally:
            db.close()

    def _save_draft_with_charts(self, client, project_id: int, headers, *, chart_ids: list) -> None:
        _save_draft(client, project_id, headers, summary="PDF fallback summary.")
        db = TestingSessionLocal()
        try:
            draft = (
                db.query(ReportDraft)
                .filter(ReportDraft.project_id == project_id)
                .order_by(ReportDraft.created_at.desc())
                .first()
            )
            assert draft is not None
            draft.selected_chart_ids_json = json.dumps(chart_ids)
            db.commit()
        finally:
            db.close()

    def test_fallback_markup_present_when_charts_selected(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        self._save_draft_with_charts(client, pid, consultant_auth_headers, chart_ids=["pdf_chart_a"])

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        assert "chart-pdf-fallback" in r.text

    def test_fallback_includes_chart_title_type_id(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        self._save_draft_with_charts(client, pid, consultant_auth_headers, chart_ids=["pdf_chart_a"])

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        body = r.text
        assert "PDF_CHART_A_TITLE" in body
        assert "BAR" in body
        assert "pdf_chart_a" in body

    def test_fallback_excludes_unselected_chart(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        # Only select chart_a; chart_b must not appear
        self._save_draft_with_charts(client, pid, consultant_auth_headers, chart_ids=["pdf_chart_a"])

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        assert "PDF_CHART_A_TITLE" in r.text
        assert "PDF_CHART_B_TITLE" not in r.text

    def test_no_fallback_when_no_charts_selected(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        self._save_draft_with_charts(client, pid, consultant_auth_headers, chart_ids=[])

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        # The CSS class appears in the <style> block, but no actual fallback div
        # should be rendered when no charts are selected.
        assert 'class="chart-pdf-fallback"' not in r.text

    def test_fallback_note_text_present(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        self._seed_run_with_charts(pid)
        self._save_draft_with_charts(client, pid, consultant_auth_headers, chart_ids=["pdf_chart_a"])

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        assert "PDF export includes chart metadata" in r.text


# ── End-to-end parity: full Report Builder → export chain ────────────────────

class TestReportBuilderExportParity:
    """Single regression checkpoint: seed a run with insights + charts, apply a
    draft that selects a subset of each, then verify every export surface
    (draft API, preview, HTML, XLSX) reflects exactly those selections.
    """

    # ── fixtures ─────────────────────────────────────────────────────────────

    def _seed_parity_run(self, project_id: int) -> int:
        db = TestingSessionLocal()
        try:
            result = {
                "project_id": project_id,
                "narrative": "AUTO_NARRATIVE_TOKEN",
                "insight_results": [
                    {
                        "insight_id": "par_ins_keep",
                        "title": "PAR_INSIGHT_KEEP",
                        "explanation": "Keep this finding.",
                        "severity": "high",
                        "recommendation": "Act on it.",
                    },
                    {
                        "insight_id": "par_ins_drop",
                        "title": "PAR_INSIGHT_DROP",
                        "explanation": "Drop this finding.",
                        "severity": "medium",
                        "recommendation": "Ignore.",
                    },
                ],
                "health_score": {"total": 72, "breakdown": {"Completeness": 72}},
                "dataset_summary": {
                    "rows": 20,
                    "columns": 4,
                    "numeric_cols": 3,
                    "categorical_cols": 1,
                    "missing_pct": 2.5,
                },
                "cleaning_report": [],
                "profile": [],
                "charts": [
                    {
                        "chart_id": "par_chart_keep",
                        "title": "PAR_CHART_KEEP_TITLE",
                        "type": "bar",
                        "chart_type": "bar",
                        "data": {
                            "labels": ["A", "B"],
                            "datasets": [{"label": "s", "data": [10, 20]}],
                        },
                    },
                    {
                        "chart_id": "par_chart_drop",
                        "title": "PAR_CHART_DROP_TITLE",
                        "type": "line",
                        "chart_type": "line",
                        "data": {
                            "labels": ["C"],
                            "datasets": [{"label": "t", "data": [5]}],
                        },
                    },
                ],
            }
            run = AnalysisResult(
                project_id=project_id,
                file_hash="parity-hash",
                result_json=json.dumps(result),
                status="report_ready",
            )
            db.add(run)
            db.commit()
            db.refresh(run)
            return run.id
        finally:
            db.close()

    def _apply_parity_draft(
        self,
        client,
        project_id: int,
        run_id: int,
        headers,
    ) -> None:
        """Save a draft with an edited summary, one selected insight, one chart."""
        _save_draft(
            client, project_id, headers,
            summary="PAR_EDITED_SUMMARY_TOKEN",
            selected=["par_ins_keep"],
        )
        _link_draft_to_run(project_id, run_id)
        db = TestingSessionLocal()
        try:
            draft = (
                db.query(ReportDraft)
                .filter(ReportDraft.project_id == project_id)
                .order_by(ReportDraft.created_at.desc())
                .first()
            )
            assert draft is not None
            draft.selected_chart_ids_json = json.dumps(["par_chart_keep"])
            db.commit()
        finally:
            db.close()

    # ── draft API parity ──────────────────────────────────────────────────────

    def test_draft_api_exposes_selected_insight_and_chart_ids(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        run_id = self._seed_parity_run(pid)
        self._apply_parity_draft(client, pid, run_id, consultant_auth_headers)

        r = client.get(f"/reports/draft/{pid}", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        body = r.json()
        assert "par_ins_keep" in (body.get("selected_insight_ids") or [])
        assert "par_chart_keep" in (body.get("selected_chart_ids") or [])

    def test_draft_api_report_result_included_insights_and_charts(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        run_id = self._seed_parity_run(pid)
        self._apply_parity_draft(client, pid, run_id, consultant_auth_headers)

        r = client.get(f"/reports/draft/{pid}", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        body = r.json()
        rr = body.get("report_result", {})

        insight_ids = [i.get("insight_id") for i in (rr.get("included_insights") or [])]
        assert "par_ins_keep" in insight_ids
        assert "par_ins_drop" not in insight_ids

        chart_ids = [c.get("chart_id") for c in (rr.get("included_charts") or [])]
        assert "par_chart_keep" in chart_ids
        assert "par_chart_drop" not in chart_ids

    # ── preview parity ────────────────────────────────────────────────────────

    def test_preview_returns_edited_summary_and_selected_insights(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        run_id = self._seed_parity_run(pid)
        self._apply_parity_draft(client, pid, run_id, consultant_auth_headers)

        r = client.get(f"/reports/preview/{pid}", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body.get("narrative") == "PAR_EDITED_SUMMARY_TOKEN"
        titles = [i.get("title") for i in (body.get("insight_results") or [])]
        assert "PAR_INSIGHT_KEEP" in titles
        assert "PAR_INSIGHT_DROP" not in titles

    # ── HTML export parity ────────────────────────────────────────────────────

    def test_html_export_full_chain(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        run_id = self._seed_parity_run(pid)
        self._apply_parity_draft(client, pid, run_id, consultant_auth_headers)

        r = client.get(f"/reports/export/{pid}?format=html", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        body = r.text

        # Edited summary
        assert "PAR_EDITED_SUMMARY_TOKEN" in body
        # Selected insight present; dropped insight absent
        assert "PAR_INSIGHT_KEEP" in body
        assert "PAR_INSIGHT_DROP" not in body
        # Chart Gallery section
        assert "Chart Gallery" in body
        assert "PAR_CHART_KEEP_TITLE" in body
        assert "PAR_CHART_DROP_TITLE" not in body
        # PDF fallback metadata block present
        assert 'class="chart-pdf-fallback"' in body
        assert "PDF export includes chart metadata" in body

    # ── Excel export parity ───────────────────────────────────────────────────

    def test_xlsx_export_full_chain(
        self, client, uploaded_project, consultant_auth_headers
    ):
        pid = uploaded_project["id"]
        run_id = self._seed_parity_run(pid)
        self._apply_parity_draft(client, pid, run_id, consultant_auth_headers)

        r = client.get(f"/reports/export/{pid}?format=xlsx", headers=consultant_auth_headers)
        assert r.status_code == 200, r.text
        wb = load_workbook(io.BytesIO(r.content), data_only=True)

        # Summary sheet carries the edited executive summary
        summary_text = "\n".join(
            str(c.value) for row in wb["Summary"].iter_rows() for c in row if c.value
        )
        assert "PAR_EDITED_SUMMARY_TOKEN" in summary_text

        # Insights sheet: selected present, dropped absent
        insights_text = "\n".join(
            str(c.value) for row in wb["Insights"].iter_rows() for c in row if c.value
        )
        assert "PAR_INSIGHT_KEEP" in insights_text
        assert "PAR_INSIGHT_DROP" not in insights_text

        # Selected Charts sheet: selected chart present, dropped chart absent
        assert "Selected Charts" in wb.sheetnames
        charts_text = "\n".join(
            str(c.value) for row in wb["Selected Charts"].iter_rows() for c in row if c.value
        )
        assert "PAR_CHART_KEEP_TITLE" in charts_text
        assert "PAR_CHART_DROP_TITLE" not in charts_text
