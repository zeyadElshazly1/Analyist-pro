"""
Excel report generator.

generate_excel_report(df, analysis_result, project_name) → bytes

Sheets:
  1. Summary        — dataset stats + health score
  2. Insights       — all insights as a table
  3. Column Profiles — per-column statistics + conditional formatting on Missing %
  4. Cleaning Report — cleaning steps
  5. Data Preview   — first 200 rows (vectorized, no iterrows)
"""
from __future__ import annotations

import io
from datetime import datetime

import pandas as pd


def generate_excel_report(
    df: pd.DataFrame,
    analysis_result: dict,
    project_name: str = "Dataset Analysis",
) -> bytes:
    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.formatting.rule import ColorScaleRule  # noqa: PLC0415
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: PLC0415
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    DARK_FILL   = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="F8FAFC", size=11)
    TITLE_FONT  = Font(bold=True, color="F8FAFC", size=14)
    THIN_BORDER = Border(bottom=Side(style="thin", color="334155"))

    def _style_header(ws, row_num: int, ncols: int) -> None:
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill      = DARK_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = THIN_BORDER

    def _auto_width(ws, min_w: int = 10, max_w: int = 50) -> None:
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            vals   = [str(c.value or "") for c in col if c.value is not None]
            width  = max(min(max(len(v) for v in vals) + 4, max_w), min_w) if vals else min_w
            ws.column_dimensions[letter].width = width

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False

    ws1["A1"] = project_name
    ws1["A1"].font = TITLE_FONT
    ws1["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1["A2"].font = Font(color="64748B", size=10, italic=True)

    summary = analysis_result.get("dataset_summary", {})
    health  = analysis_result.get("health_score", {})
    ws1.append([])
    ws1.append(["Metric", "Value"])
    _style_header(ws1, ws1.max_row, 2)

    for label, val in [
        ("Rows",                 summary.get("rows", "—")),
        ("Columns",              summary.get("columns", "—")),
        ("Numeric Columns",      summary.get("numeric_cols", "—")),
        ("Categorical Columns",  summary.get("categorical_cols", "—")),
        ("Missing Data %",       f"{summary.get('missing_pct', 0):.1f}%"),
        ("Health Score",         f"{health.get('total', health.get('overall', '—'))}/100"),
    ]:
        ws1.append([label, val])

    ws1.freeze_panes = "A5"
    _auto_width(ws1)

    # ── Sheet 2: Insights ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Insights")
    ws2.sheet_view.showGridLines = False

    headers2 = ["#", "Type", "Severity", "Finding", "Evidence", "Action"]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))
    ws2.freeze_panes = "A2"

    for i, ins in enumerate(analysis_result.get("insights", []), 1):
        ws2.append([
            i,
            ins.get("type", ""),
            ins.get("severity", ""),
            ins.get("finding", ins.get("title", "")),
            ins.get("evidence", ""),
            ins.get("action", ""),
        ])
    _auto_width(ws2, max_w=60)

    # ── Sheet 3: Column Profiles (+ conditional formatting) ───────────────────
    ws3 = wb.create_sheet("Column Profiles")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Column", "Type", "Missing %", "Unique", "Mean", "Std", "Min", "Max"]
    ws3.append(headers3)
    _style_header(ws3, 1, len(headers3))
    ws3.freeze_panes = "A2"

    profile = analysis_result.get("profile", [])
    if isinstance(profile, dict):
        profile_cols = profile.get("columns", [])
    else:
        profile_cols = profile if isinstance(profile, list) else []
    for col in profile_cols:
        ws3.append([
            col.get("name") or col.get("column", ""),
            col.get("dtype", ""),
            round(float(col.get("missing_pct", 0) or 0), 1),
            col.get("unique_count", col.get("n_unique", "—")),
            round(float(col.get("mean") or 0), 4) if col.get("mean") is not None else "—",
            round(float(col.get("std")  or 0), 4) if col.get("std")  is not None else "—",
            col.get("min", "—"),
            col.get("max", "—"),
        ])

    if ws3.max_row > 1:
        ws3.conditional_formatting.add(
            f"C2:C{ws3.max_row}",
            ColorScaleRule(
                start_type="num", start_value=0,   start_color="10B981",
                mid_type="num",   mid_value=15,    mid_color="F59E0B",
                end_type="num",   end_value=100,   end_color="EF4444",
            ),
        )
    _auto_width(ws3)

    # ── Sheet 4: Cleaning Report ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Cleaning Report")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Step", "Detail", "Impact"]
    ws4.append(headers4)
    _style_header(ws4, 1, len(headers4))
    ws4.freeze_panes = "A2"

    for item in analysis_result.get("cleaning_report", []):
        if isinstance(item, dict):
            ws4.append([item.get("step", ""), item.get("detail", ""), item.get("impact", "")])
        else:
            ws4.append([str(item), "", ""])
    _auto_width(ws4, max_w=60)

    # ── Sheet 5: Data Preview (vectorized — no iterrows) ──────────────────────
    if not df.empty:
        ws5 = wb.create_sheet("Data Preview")
        ws5.sheet_view.showGridLines = False

        preview = df.head(200)
        ws5.append(list(preview.columns))
        _style_header(ws5, 1, len(preview.columns))
        ws5.freeze_panes = "A2"

        for row_vals in preview.to_numpy():
            ws5.append([
                None if (isinstance(v, float) and v != v)  # NaN → None
                else v if isinstance(v, (int, float, type(None)))
                else str(v)
                for v in row_vals
            ])
        _auto_width(ws5)

    # ── Sheet 6: Report Info (trust labels) ──────────────────────────────────
    ws6 = wb.create_sheet("Report Info")
    ws6.sheet_view.showGridLines = False

    ws6["A1"] = "Report Metadata"
    ws6["A1"].font = TITLE_FONT

    cleaning_steps = len(analysis_result.get("cleaning_report", []))
    n_rows, n_cols = df.shape
    health = analysis_result.get("health_score", {})
    health_total = health.get("total", health.get("overall", "—")) if isinstance(health, dict) else health

    meta_rows = [
        ("Report title", project_name),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M UTC")),
        ("Rows analyzed", f"{n_rows:,}"),
        ("Columns analyzed", n_cols),
        ("Health score", f"{health_total}/100"),
        ("Cleaning steps applied", cleaning_steps),
        ("Source", "Analyst Pro — analyst-pro.io"),
        ("Note", "Narrative sections marked 'AI-generated' were produced by Claude AI and should be reviewed."),
    ]

    ws6.append([])
    for label, val in meta_rows:
        ws6.append([label, str(val)])
    _auto_width(ws6)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
